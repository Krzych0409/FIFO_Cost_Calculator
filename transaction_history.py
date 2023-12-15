import openpyxl
from colorama import init, Fore, Back
import pandas as pd


class Data:
    def get_launch_data(self, wb_source):
        """Import initial configuration data from an Excel file."""

        # Load configuration sheet
        try:
            self.ws_launch_data = wb_source['launch_data']
        except KeyError as err:
            print(f'KeyError: {err}')
            return False

        # Load PURCHASES option
        self.sheet_name_p = str(self.ws_launch_data['D11'].value)
        self.first_row_p = int(self.ws_launch_data['D12'].value)
        self.col_date_p = int(self.ws_launch_data['D13'].value)
        self.col_name_p = int(self.ws_launch_data['D14'].value)
        self.col_ppi_p = int(self.ws_launch_data['D15'].value)
        self.col_volume_p = int(self.ws_launch_data['D16'].value)
        self.col_unr_volume_p = int(self.ws_launch_data['D17'].value)
        # self.min_col_p = min(self.col_date_p, self.col_name_p, self.col_ppi_p, self.col_volume_p, self.col_unr_volume_p)
        # self.max_col_p = max(self.col_date_p, self.col_name_p, self.col_ppi_p, self.col_volume_p, self.col_unr_volume_p)

        # Load SALES option
        self.sheet_name_s = str(self.ws_launch_data['G11'].value)
        self.first_row_s = int(self.ws_launch_data['G12'].value)
        self.col_date_s = int(self.ws_launch_data['G13'].value)
        self.col_name_s = int(self.ws_launch_data['G14'].value)
        self.col_ppi_s = int(self.ws_launch_data['G15'].value)
        self.col_volume_s = int(self.ws_launch_data['G16'].value)
        self.col_cogs_s = int(self.ws_launch_data['G17'].value)
        # self.min_col_s = min(self.col_date_s, self.col_name_s, self.col_ppi_s, self.col_volume_s, self.col_cogs_s)
        # self.max_col_s = max(self.col_date_s, self.col_name_s, self.col_ppi_s, self.col_volume_s, self.col_cogs_s)

        return True


class Transaction:
    def __init__(self, date, name, price_per_item, volume):
        self.date = date
        self.name = name
        self.price_per_item = price_per_item
        self.volume = volume

class Purchase(Transaction):
    all_purchases = []

    def __init__(self, date, name, price_per_item, volume, unrealized_volume):
        super().__init__(date, name, price_per_item, volume)
        self.unrealized_volume = unrealized_volume

        Purchase.all_purchases.append(self)

class Sale(Transaction):
    all_sales = []

    def __init__(self, date, name, price_per_item, volume, cost_of_goods_sold):
        super().__init__(date, name, price_per_item, volume)
        self.cost_of_goods_sold = cost_of_goods_sold

        Sale.all_sales.append(self)

class Portfolio:
    def __init__(self):
        self.transactions = []

    def add_transaction(self, transaction):
        self.transactions.append(transaction)


def main():
    # name your excel history
    excel_name = 'source.xlsx'

    # Colorama
    init(autoreset=True)

    # Load excel file
    wb_source = openpyxl.load_workbook(excel_name, data_only=False)

    # If configuration data not load end program
    data = Data()
    if data.get_launch_data(wb_source) == False:
        return False

    # Load worksheet sales
    ws_s = wb_source[data.sheet_name_s]

    # Do DataFrame with sales
    df_s = pd.read_excel(excel_name, sheet_name='sales', header=data.first_row_s-2)
    print(df_s)
    print()

    # Loop iterates through a list of sales transactions
    for row_s in df_s.itertuples():
        # Check if the row has a name and if the 'cogs' value has not been entered
        if pd.isna(row_s[data.col_name_s]) or not pd.isna(row_s[data.col_cogs_s]):
            continue

        print(f'Start sale: {row_s[data.col_name_s]} - wolumen {row_s[data.col_volume_s]} - ppi {row_s[data.col_ppi_s]}')

        # Do DataFrame with purchases
        df_p = pd.read_excel(excel_name, sheet_name='purchases', header=data.first_row_p-2)

        # Load worksheet purchases
        ws_p = wb_source[data.sheet_name_p]

        # Variables needed to calculate "cogs"
        required_volume = float(row_s[data.col_volume_s])
        missing_volume = required_volume
        found_volume = 0.0
        total_cost = 0.0
        changes_unr = []

        # Loop iterates through a list of purchases transactions
        for row_p in df_p.itertuples():
            # Searching for a pair. Comparison of the name from the sale transaction with the name from the purchase transaction.
            if row_s[data.col_name_s].lower().strip() != row_p[data.col_name_p].lower().strip():
                continue

            # Check if there is available "unrealized volume"
            if not row_p[data.col_unr_volume_p] > 0:
                continue

            # Encountering a purchase transaction that followed a sale transaction
            if row_s[data.col_date_s] < row_p[data.col_date_p]:
                print(f'{Fore.YELLOW}The date of the sale transaction is later than the purchase transaction:'
                      f'\n{row_s[data.col_name_s]} : {row_s[data.col_volume_s]} : {row_s[data.col_ppi_s]} - still searching')
                continue

            # New value "unrealized volume"
            new_unr_volume = float(row_p[data.col_unr_volume_p])
            print(f'old unrealized volume: {new_unr_volume} - must be float with ".0"')

            # If "unrealized volume" is greater than "missing_volume"
            if row_p[data.col_unr_volume_p] >= missing_volume:
                new_unr_volume -= missing_volume  # new_unr_volume >= 0
                found_volume += missing_volume
                total_cost += float(row_p[data.col_ppi_p]) * missing_volume
                print(f'total_cost += {float(row_p[data.col_ppi_p])} * {missing_volume} = {float(row_p[data.col_ppi_p]) * missing_volume}')
                missing_volume -= missing_volume  # missing_volume = 0

            # If "missing_volume" is greater than "unrealized volume"
            else:
                missing_volume -= new_unr_volume
                found_volume += new_unr_volume
                total_cost += float(row_p[data.col_ppi_p]) * new_unr_volume
                print(f'total_cost += {float(row_p[data.col_ppi_p])} * {new_unr_volume} = {float(row_p[data.col_ppi_p]) * new_unr_volume}')
                new_unr_volume -= new_unr_volume  # new_unr_volume = 0

            print(f'Unrealized volume (row_p = {row_p.Index + data.first_row_p}, column_p = {data.col_unr_volume_p}) = {new_unr_volume}')

            # Enter tuple to list ("unrealized volume", number of row, number of column)
            changes_unr.append((new_unr_volume, row_p.Index + data.first_row_p, data.col_unr_volume_p))

            if missing_volume == 0 and found_volume == required_volume:
                print(f'Total cost (row_s = {row_s.Index + data.first_row_s}, column_s = {data.col_cogs_s}) = {total_cost}')
                print(f'End sale: {row_s[data.col_name_s]} - wolumen {row_s[data.col_volume_s]} - ppi {row_s[data.col_ppi_s]}')

                # Enter all "unrealized volume" in purchases sheet
                for unr in changes_unr:
                    # Enter new "unrealized volume" in purchase
                    ws_p.cell(row=unr[1], column=unr[2]).value = unr[0]

                # Enter new "cogs" in sale
                ws_s.cell(row=row_s.Index + data.first_row_s, column=data.col_cogs_s).value = total_cost

                # Save changes ("unrealized volume" and "cogs")
                try:
                    wb_source.save(excel_name)
                except PermissionError as err:
                    print(f'{Fore.RED}Unsaved data in excel file!\nMake sure the excel file is closed. Error details: \n{err}')

                print()
                print()
                break

        # Needed volume was not found in the purchase transactions
        else:
            print(f'{Fore.RED}Error: Not found all required purchase transactions!\n'
                  f'Sale row = {row_s.Index + data.first_row_s} : {row_s[data.col_name_s]}\n')


if __name__ == '__main__':
    main()




